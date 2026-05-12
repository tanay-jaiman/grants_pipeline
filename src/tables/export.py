#!/usr/bin/env python3

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SAGE_GREEN = "A8C69B"
LIGHT_GRAY = "EDEDED"
WHITE = "FFFFFF"
BLACK = "000000"
TABLE_PADDING_ROWS = 2
TABLE_PADDING_COLS = 1


LAYOUT = {
    "stats": {"anchor": "top", "row": 4, "col": 7},
    "unique": {"below": "stats", "col": 7},
    "range": {"anchor": "top", "row": 3, "col": 10},
    "category": {"below": "range", "col": 10},
    "states": {"below": "category", "col": 10},
    "location": {"anchor": "top", "row": 3, "after": "range"}
}


def export_year_sheet(writer, sheet_name, tables):
    layout_positions = _calculate_layout(tables)

    for table_index, table in enumerate(tables, start=1):
        config = table["config"]
        dataframe = table["data"]
        title = table["title"]
        start_row, start_col = layout_positions.get(
            table.get("key"),
            (config["start_row"], config["start_col"])
        )
        show_title = table.get("show_title", True)

        if show_title:
            worksheet = writer.book[sheet_name]
            title_cell = worksheet.cell(
                row=start_row + 1,
                column=start_col + 1
            )
            title_cell.value = title.upper()
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="left")
            data_start_row = start_row + 1
        else:
            data_start_row = start_row

        dataframe.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=data_start_row,
            startcol=start_col,
            index=False
        )

        worksheet = writer.book[sheet_name]
        header_row = data_start_row + 1
        last_row = header_row + len(dataframe)
        last_col = start_col + len(dataframe.columns)
        _format_table(
            worksheet=worksheet,
            dataframe=dataframe,
            header_row=header_row,
            start_col=start_col + 1,
            last_row=last_row,
            last_col=last_col,
            header_fill=SAGE_GREEN,
            header_font_color=BLACK,
            band_fill=table.get("band_fill", LIGHT_GRAY)
        )

    _format_sheet(writer.book[sheet_name])


def _calculate_layout(tables):
    positions = {}
    bounds = {}

    for table in tables:
        key = table.get("key")

        if not key:
            continue

        rule = LAYOUT.get(key, {})
        config = table["config"]
        show_title = table.get("show_title", True)
        height = len(table["data"]) + 1 + (1 if show_title else 0)
        width = len(table["data"].columns)

        if not rule:
            start_row = config["start_row"]
        elif rule.get("anchor") == "top":
            start_row = rule.get("row", config["start_row"])
        else:
            previous = bounds[rule["below"]]
            start_row = previous["end_row"] + TABLE_PADDING_ROWS + 1

        if "after" in rule:
            previous = bounds[rule["after"]]
            start_col = previous["end_col"] + TABLE_PADDING_COLS + 1
        else:
            start_col = rule.get("col", config["start_col"])

        positions[key] = (start_row, start_col)
        bounds[key] = {
            "end_row": start_row + height - 1,
            "end_col": start_col + width - 1
        }

    return positions


def _format_table(
    worksheet,
    dataframe,
    header_row,
    start_col,
    last_row,
    last_col,
    header_fill,
    header_font_color,
    band_fill
):
    thin_border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7")
    )

    for col in range(start_col, last_col + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(bold=True, color=header_font_color, size=11)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = thin_border

    for row in range(header_row + 1, last_row + 1):
        is_total = worksheet.cell(row=row, column=start_col).value == "TOTAL"

        for col in range(start_col, last_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            if is_total:
                cell.font = Font(bold=True, color=BLACK)
                cell.fill = PatternFill("solid", fgColor=SAGE_GREEN)
            elif (row - header_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=band_fill)
            else:
                cell.fill = PatternFill("solid", fgColor=WHITE)

            header = worksheet.cell(row=header_row, column=col).value
            _apply_number_format(cell, header)

    worksheet.row_dimensions[header_row].height = 42


def _apply_number_format(cell, header):
    header_text = str(header or "").lower()

    if (
        "amount" in header_text
        or header_text in {"value"}
    ) and isinstance(cell.value, (int, float)):
        cell.number_format = '$#,##0'

    if (
        "percent" in header_text
        or "%" in header_text
    ) and isinstance(cell.value, (int, float)):
        cell.number_format = '0.00"%"'


def _format_sheet(worksheet):
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True

    def set_widths(start_col, widths):
        for offset, width in enumerate(widths):
            column = get_column_letter(start_col + offset)
            worksheet.column_dimensions[column].width = width

    set_widths(1, [30, 34, 26, 24, 38, 14])
    set_widths(8, [16, 16])
    set_widths(11, [18, 36, 14, 15, 18])
    set_widths(17, [24, 14, 20, 22, 22])

    worksheet.sheet_format.defaultRowHeight = 22
